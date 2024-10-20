from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


class ExcelProcessorMontrekFormatter:
    @staticmethod
    def format_excel(writer, sheet_name="Sheet1"):
        # Access the workbook and sheet
        worksheet = writer.sheets[sheet_name]

        # Define the styles
        header_fill = PatternFill(
            start_color="004767", end_color="004767", fill_type="solid"
        )  # Dark blue background
        header_font = Font(color="FFFFFF", bold=True)  # White, bold text

        even_row_fill = PatternFill(
            start_color="E6F2F8", end_color="E6F2F8", fill_type="solid"
        )  # Pale blue background
        odd_row_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )  # White background

        thin_border = Border(
            bottom=Side(style="thin", color="E0E0E0")
        )  # Light bottom border

        # Get the range of data
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            for cell in row:
                # Apply styles to header row
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left")
                    cell.border = thin_border
                else:
                    # Apply alternating row styles
                    if row_idx % 2 == 0:  # Even row
                        cell.fill = even_row_fill
                    else:  # Odd row
                        cell.fill = odd_row_fill
                    # Apply other cell styles
                    cell.border = thin_border
                    # Format floats to ",.2f"
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.00"  # Format as ",.2f"
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
        # Adjust column widths to fit content
        for col in worksheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get column name
            for cell in col:
                try:
                    # Get the length of the cell value (converted to string)
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Set the column width
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width
